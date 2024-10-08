from PIL import Image, ImageDraw, ImageFont
import textwrap
import os
import openpyxl
import re

# 哈哈哈哈哈哈

def generate_title_image(output_folder, title_lines, content, img_title_path, font_title_path, font_title_size,
                         anchor_title_y, line_spacing, content_params, title_color):
    #去重，把正文中与标题一样的第一段剔除
    content = preprocess_content(title_lines, content)

    # 打开指定路径的图片作为标题背景
    img_title = Image.open(img_title_path)
    # 使用指定的字体文件和大小创建字体对象
    font_title = ImageFont.truetype(font_title_path, font_title_size)
    # 在图片上创建绘图对象
    draw = ImageDraw.Draw(img_title)
    # 获取图片宽度
    img_title_width = img_title.width

    # 绘制标题
    width_char = 15  #代表多少个字符换行

    # 使用textwrap将标题文本按指定宽度换行
    lines = textwrap.wrap(title_lines, width= width_char)

    # 从content_params中获取左边距
    left_margin = content_params["left_margin"]

    # 逐行绘制标题文本
    for line in lines:
        draw.text((left_margin, anchor_title_y), line, fill=title_color, font=font_title)
        # 更新下一行的y坐标
        anchor_title_y += font_title_size + line_spacing


    # 绘制内容
    content_y = anchor_title_y
    content_y += content_params['font_size'] + content_params['line_spacing'] * 0.6


    # 调用draw_content_with_special_chars函数绘制内容，并获取可能未绘制的剩余内容
    remaining_content = draw_content_with_special_chars(draw, content, content_y, content_params, img_title.width,
                                                        img_title.height)

    # 保存生成的图片
    img_title.save(os.path.join(output_folder, f"{title_lines}-1.jpg"))
    print(f'标题图片生成完毕')
    # 返回剩余未绘制的内容
    return remaining_content
def preprocess_content(title, content):
    lines = content.split('\n')
    if lines and lines[0].strip() == title.strip():
        return '\n'.join(lines[1:]).strip()
    return content

def draw_content_with_special_chars(draw, content, start_y, params, img_width, max_height):
    font = ImageFont.truetype(params['font_path'], params['font_size'])
    bold_font = ImageFont.truetype('SourceHanSansCN-Medium.otf', params['font_size'])

    y_position = start_y
    lines = content.split('\n')
    remaining_lines = []
    left_margin = params['left_margin']
    max_width = img_width - left_margin - 72

    special_tags = ["网友提问：", "提问：", "王盐："]

    # 新增字间距参数
    char_spacing = params.get('char_spacing', 0)  # 默认为0，可以在params中设置

    for line in lines:
        if y_position + params['font_size'] > max_height - 100:
            remaining_lines.append(line)
            continue

        x_position = left_margin

        special_start = None
        for tag in special_tags:
            if line.strip().startswith(tag):
                special_start = tag
                break

        chars = list(line)
        i = 0

        while i < len(chars):
            char = chars[i]
            if special_start and i < len(special_start):
                # 特殊字符处理
                char_width = draw.textlength(char, font=bold_font)
                if x_position + char_width > img_width - 72:
                    y_position += params['font_size'] + params['line_spacing']
                    x_position = left_margin
                    if y_position + params['font_size'] > max_height - 100:
                        remaining_lines.append(''.join(chars[i:]))
                        break
                draw.text((x_position, y_position), char, fill=params['special_color'], font=bold_font)
                x_position += char_width + char_spacing  # 添加字间距
            else:
                # 普通字符处理
                char_width = draw.textlength(char, font=font)
                if x_position + char_width > img_width - 72:
                    y_position += params['font_size'] + params['line_spacing']
                    x_position = left_margin
                    if y_position + params['font_size'] > max_height - 100:
                        remaining_lines.append(''.join(chars[i:]))
                        break
                draw.text((x_position, y_position), char, fill=params['text_color'], font=font)
                x_position += char_width + char_spacing  # 添加字间距

            i += 1

        y_position += params['font_size'] + params['line_spacing'] * 0.6

    return '\n'.join(remaining_lines)

def generate_content_new_images(output_folasder, title_lines, content, params):
    img_template = Image.open(params['img_path'])
    img_counter = 2

    while content.strip():
        img = img_template.copy()
        draw = ImageDraw.Draw(img)
        content = draw_content_with_special_chars(draw, content, params['start_y'], params, img_template.width,
                                                  img_template.height)

        img.save(os.path.join(output_folder, f"{title_lines}-{img_counter}.jpg"))
        print(f'生成了第 {img_counter} 张内容图片')
        img_counter += 1

    print(f'所有内容图片绘制完毕，共生成 {img_counter - 2} 张内容图片')
def read_excel_and_print(file_path, output_folder, title_params, content_params):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=2):
        column1_value = row[0].value
        column2_value = row[1].value

        print(f"处理: {column1_value}")

        remaining_content = generate_title_image(output_folder, column1_value, column2_value, **title_params,
                                                 content_params=content_params)


        if remaining_content.strip():

            generate_content_new_images(output_folder, column1_value, remaining_content, content_params)

        print("---" * 20)

    workbook.close()


# 主程序
if __name__ == "__main__":

    # 标题基础参数设置
    title_params = {
        "img_title_path": os.path.abspath('template-background.jpg'),
        "font_title_path": os.path.abspath('SourceHanSansCN-Medium.otf'),
        "font_title_size": 70,
        "anchor_title_y": 50,
        "line_spacing": 30,#行间距
        "title_color": (190, 10, 10),   #(128, 30, 63)
    }

    # 内容基础参数设置
    img_template = Image.open(os.path.abspath('template-new-content.jpg'))
    img_width = img_template.width
    content_params = {
        "img_path": os.path.abspath('template-background.jpg'),
        "font_path": os.path.abspath('SourceHanSansCN-Normal.otf'),
        "font_size": 52,
        "line_spacing": 30,#行间距
        "chars_per_line": 22,# 每行固定字符数
        "text_color":  (0, 0, 0),
        "special_color":  (190, 10, 10),   #(128, 30, 63)
        "start_y": 72,
        "left_margin": 72,
        "char_spacing": 4 #字间距
    }

    # 读取EXCEL文件
    file_path = '八爪鱼演示站点20230105-230106.xlsx'
    output_folder = "图片文件夹"

    # 确保输出文件夹存在
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)

    # 调用函数并传递文件路径作为参数
    print("开始处理 Excel 文件...")
    read_excel_and_print(file_path, output_folder, title_params, content_params)
    print("处理完成！")